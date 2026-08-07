from datetime import datetime

from dataprep import full_dataprep
import pandas as pd
from isolation_forest import run_if
from models.kittyboost import train_catboost
from querystep import uncertainty_query, novelty_scores
from greedy import greedy_iteration, new_state
from openpyxl import Workbook, load_workbook

def test_logger(header, results):
    try:
        wb = load_workbook("test_results.xlsx")
        ws = wb.active
    except:
        wb = Workbook()
        ws = wb.active
    
    ws.append([header])
    for result_str in results:
        ws.append([result_str])
    ws.append(["'" + "=" * 80])  # separator line
    
    wb.save("test_results.xlsx")
    print("Results saved to test_results.xlsx")

def retrain_catboost(df, l=10, corrected_weights=100, corrected_saved=True, strategy="entropy",
                     greedy_batching=False, greedy_T=None, compute_shap=False, skip_retrain_on_skip=False):
    """Retrain CatBoost iteratively by correcting samples based on the specified strategy.

    Greedy-Modus (greedy_batching=True): pro Iteration genau 1 HITL-Review (most uncertain
    Case M). Sein Label wird via globaler Voronoi-Partition auf alle Cases im Radius T
    propagiert; Re-Selection mit Labeländerung splittet das Cluster (neues Center).

    Args:
        df (pd.DataFrame): Input DataFrame mit 'label', 'pred_label', 'posting_id'.
        l (int): Anzahl Iterationen (= Anzahl HITL-Reviews im Greedy-Modus).
        corrected_weights (int): Gewicht korrigierter Samples beim Retrain.
        corrected_saved (bool): Geprüfte posting_ids in zukünftigen Queries ausschließen.
        strategy (str): 'entropy', 'margin', 'novelty'.
        greedy_batching (bool): Greedy Cluster Correction aktivieren.
        greedy_T (float): Cluster-Radius (None -> wird geschätzt).
        compute_shap (bool): SHAP-Werte pro Trainingslauf berechnen.
        skip_retrain_on_skip (bool): Bei greedy 'skip'-Iterationen das (ergebnisgleiche)
            Retrain überspringen.
    """
    if greedy_batching:
        state = new_state(greedy_T)
    else:
        state = {"corrected": []}

    novelty_cache = novelty_scores(df) if strategy == "novelty" else None

    results = []
    target_iters = {1, l//2 + 1, l}
    cat_importances = None
    precision = recall = 0.0
    tn = fp = fn = tp = 0

    for i in range(l):
        print(f"\n=== Iteration {i+1} ===")

        if greedy_batching:
            df, state, meta = greedy_iteration(df, strategy, state)
            if meta.get("type") == "no_candidates":
                print("No more candidates in pool.")
                break
            corrected_ids = list(state["directly_corrected"] | state["covered"])
            skip_retrain = skip_retrain_on_skip and meta.get("type") == "skip"
        else:
            uncertain_df = uncertainty_query(df, strategy, exclude_posting_ids=state["corrected"],
                                             novelty_scores=novelty_cache)
            meta = {}
            mask = df['posting_id'].isin(uncertain_df['posting_id'])
            meta["misspredicted"] = (df.loc[mask, 'label'] != df.loc[mask, 'pred_label']).sum()
            df.loc[mask, 'pred_label'] = df.loc[mask, 'label']
            if corrected_saved:
                state["corrected"].extend(uncertain_df['posting_id'].tolist())
            corrected_ids = state["corrected"]
            skip_retrain = False

        if skip_retrain:
            print(f"Skip-Retrain (keine Labeländerung) | "
                  f"Precision: {precision:.4f}, Recall: {recall:.4f}")
        else:
            df, cat_importances, precision, recall, cat_model, tn, fp, fn, tp = train_catboost(
                df, corrected_ids=corrected_ids, corrected_weights=corrected_weights,
                compute_shap=compute_shap
            )
            print(f"Precision: {precision:.4f}, Recall: {recall:.4f}")
            if greedy_batching:
                print(f"type={meta.get('type')}, centers={meta.get('n_centers')}, "
                      f"covered={meta.get('n_covered')}, flipped={meta.get('n_flipped')}")
            else:
                print(f"Misspredicted: {meta.get('misspredicted')}")

        if (i+1) in target_iters:
            base = (f"Iteration: {i+1}, Precision: {precision:.4f}, Recall: {recall:.4f}, "
                    f"TP: {tp}, FP: {fp}, TN: {tn}, FN: {fn}")
            if skip_retrain:
                extra = ", type=skip (retrain skipped)"
            elif greedy_batching:
                extra = (f", type={meta.get('type')}, centers={meta.get('n_centers')}, "
                         f"covered={meta.get('n_covered')}, cum_direct={meta.get('cumulative_direct')}, "
                         f"flipped={meta.get('n_flipped')}, prop_acc={meta.get('propagation_accuracy'):.4f}, "
                         f"T={state['T']:.4f}")
            else:
                extra = f", Misspredicted: {meta.get('misspredicted')}"
            results.append(base + extra)

    return df, cat_importances, precision, recall, results

def incremental_catboost(df, l=10, return_full_data=False):
    corrected = []
    cat_model = None
    results = []
    counter = 0
    
    for i in range(l):
        print(f"\n=== Incremental Iteration {i+1} ===")
        
        # Get uncertain samples
        uncertain_df = uncertainty_query(df, strategy="entropy", exclude_posting_ids=corrected)
        
        if uncertain_df.empty:
            print("No more uncertain samples.")
            break
        
        counter += 1
        
        # Vectorized: Replace labels for all uncertain samples at once
        mask = df['posting_id'].isin(uncertain_df['posting_id'])
        misspredicted = (df.loc[mask, 'label'] != df.loc[mask, 'pred_label']).sum()
        df.loc[mask, 'pred_label'] = df.loc[mask, 'label']
        corrected.extend(uncertain_df['posting_id'].tolist())
        
        # Get the updated uncertain_df from df
        uncertain_df_updated = df[df['posting_id'].isin(uncertain_df['posting_id'])].copy()
        
        # Choose training dataset based on return_full_data flag
        train_data = df if return_full_data else uncertain_df_updated
        
        # Train incrementally
        df, cat_importances, precision, recall, cat_model, tn, fp, fn, tp = train_catboost(
            train_data, incremental=True, inc_model=cat_model, full_data=df
        )
        
        print(f"Precision: {precision:.4f}, Recall: {recall:.4f}")
        print(f"Misspredicted: {misspredicted}")
        
        if counter in {1, l//2 + 1, l}:
            results.append(f"Iteration: {counter}, Precision: {precision:.4f}, Recall: {recall:.4f}, TP: {tp}, FP: {fp}, TN: {tn}, FN: {fn}")
    
    return df, cat_importances, precision, recall, cat_model, results

def replace_posting(df, posting_id, replacement):
    df.loc[df['posting_id'] == posting_id, 'pred_label'] = replacement
    return df

def run_unsupervised():
    df, labels = full_dataprep()
    df_if = run_if(df, labels, verbose=False)
    df_cat, cat_importances, precision, recall, cat_model,tn, fp, fn, tp = train_catboost(df_if, verbose=True)
    print(f"\nInitial Precision: {precision:.4f}, Initial Recall: {recall:.4f}")
    return df_cat, cat_importances, precision, recall, cat_model

def run_supervised(training_strat= 'retrain', l=10, corrected_weights=100, corrected_saved=True, strategy="entropy", return_full_data=False, greedy_batching=False, greedy_T=None, compute_shap=False, skip_retrain_on_skip=False):
    df, cat_importances, precision, recall, cat_model = run_unsupervised()
    if training_strat == 'incremental':
        df, cat_importances, precision, recall, cat_model, results = incremental_catboost(df, l, return_full_data)
    else:
        df, cat_importances, precision, recall, results = retrain_catboost(
            df, l, corrected_weights, corrected_saved, strategy,
            greedy_batching=greedy_batching, greedy_T=greedy_T,
            compute_shap=compute_shap, skip_retrain_on_skip=skip_retrain_on_skip)
    test_logger(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] | training_strat= {training_strat}, l={l}, corrected_weights = {corrected_weights}, corrected_saved = {corrected_saved}, strategy = {strategy}, greedy_batching = {greedy_batching}, greedy_T = {greedy_T}, compute_shap = {compute_shap}, skip_retrain_on_skip = {skip_retrain_on_skip}", results)
    return df, cat_importances, precision, recall, cat_model

if __name__ == "__main__":

    df, cat_importances, precision, recall, cat_model = run_supervised(training_strat='retrain', l=50, corrected_weights=100, corrected_saved=True, strategy="entropy", return_full_data=False, greedy_batching=True, greedy_T=0.5)
    df, cat_importances, precision, recall, cat_model = run_supervised(training_strat='retrain', l=50, corrected_weights=100, corrected_saved=True, strategy="margin", return_full_data=False, greedy_batching=True, greedy_T=0.5)
    